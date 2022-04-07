import itertools

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime as dt
from ME_Data import ME, Cell
from utils import *


def generate_5g_5g_cosite(me_id, workbook=None, save_file=False):
    mes = [me.strip().upper() for me in me_id.split(";")]
    target_me, mes_data = load_input(mes, workbook)

    generate_cosites(target_me, mes_data)

    if save_file:
        if len(mes) == 1:
            target_me.save(f"data/{mes_data[0].site_name}_singleSite_Baseline_{dt.now().strftime('%d%m%Y_%H%M%S')}.xlsx")
        else:
            target_me.save(f"data/{mes_data[0].site_name}_multiSite_Baseline_{dt.now().strftime('%d%m%Y_%H%M%S')}.xlsx")
    else:
        return target_me


def load_input(me_list, target):
    mes_data = []
    if target is None:
        target_me = load_workbook("data/input.xlsx")
    else:
        target_me = target
    sheet_nameNRCell = "NRCellDU"
    try:
        proj_paramsNRCell = target_me[sheet_nameNRCell]
    except:
        raise Exception("Cannot find " + sheet_nameNRCell + " MO in input file")

    for me_id in me_list:
        me_info = ME(me_id)
        n_cells = []
        for i in range(6, get_max_row(target_me, sheet_nameNRCell) + 1):
            if proj_paramsNRCell[f"D{i}"].value == me_id:
                if not me_info.initialized:
                    me_info.subnetid = proj_paramsNRCell[f"C{i}"].value
                    me_info.gnbdu = find_cucp(proj_paramsNRCell[f"F{i}"].value)
                    me_info.site_name = proj_paramsNRCell[f"E{i}"].value
                if int(proj_paramsNRCell[f"G{i}"].value) in range(700, 710):
                    n_cells.append(proj_paramsNRCell[f"G{i}"].value)
                me_info.initialized = True
        if not hasattr(me_info, "gnbdu"):
            raise Exception("Unable to find " + me_id + " in NRCellDU")
        me_info.n_cells = n_cells
        mes_data.append(me_info)

    return target_me, mes_data


def generate_cosites(working_wb, mes_data):
    sheet = "NRCellRelation"
    sheet_data = working_wb[sheet]

    last_row = get_max_row(working_wb, sheet)

    for me in mes_data:
        reference_row = find_reference_row(sheet_data, last_row, me)
        if reference_row is None:
            continue
        combs = itertools.permutations(me.n_cells, 2)
        target_row = get_max_row(working_wb, sheet) + 1

        for comb in combs:
            rel_exists = relation_exists_5g(comb, sheet_data, last_row, me)
            if rel_exists < 0:
                copy_row(reference_row, target_row, working_wb, sheet, sheet_data)
                replace_5g_cosite_params(target_row, sheet_data, me, comb, last_row)
                add_modifier(target_row, "A", sheet_data)
                target_row += 1
            else:
                print("Relation", comb[0], "->", comb[1], "already exists")


if __name__ == "__main__":
    generate_5g_5g_cosite("PVAX0844", None, True)