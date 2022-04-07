from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime as dt
from ME_Data import ME, Cell
from utils import *


def generate_4g_4g_cosite(me_id, workbook=None, save_file=False):
    mes = [me.strip().upper() for me in me_id.split(";")]
    target_me, mes_data = load_input(mes, workbook)

    generate_cosites(target_me, mes_data)

    if save_file:
        if len(mes) == 1:
            target_me.save(f"data/{mes_data[0].site_name}_singleSite_Baseline_{dt.now().strftime('%d%m%Y_%H%M%S')}.xlsx")
        else:
            target_me.save(f"data/{mes_data[0].site_name}_multiSite_4G_4G_Baseline_{dt.now().strftime('%d%m%Y_%H%M%S')}.xlsx")
    else:
        return target_me


def load_input(me_list, target):
    mes_data = []
    if target is None:
        target_me = load_workbook("data/input.xlsx")
    else:
        target_me = target
    sheet_nameLTECell = "CUEUtranCellFDDLTE"
    try:
        proj_paramsLTECell = target_me[sheet_nameLTECell]
    except:
        raise Exception("Cannot find " + sheet_nameLTECell + " MO in input file")

    for me_id in me_list:
        me_info = ME(me_id)
        n_cells = 0
        cell_sectors = []
        for i in range(6, get_max_row(target_me, sheet_nameLTECell) + 1):
            if proj_paramsLTECell[f"D{i}"].value == me_id:
                if not me_info.initialized:
                    me_info.subnetid = proj_paramsLTECell[f"C{i}"].value
                    me_info.enbcu = find_cucp(proj_paramsLTECell[f"F{i}"].value)
                    me_info.site_name = proj_paramsLTECell[f"E{i}"].value
                    # me_info.user_label = proj_paramsLTECell[f"H{i}"].value
                    # me_info.cells = []
                if int(proj_paramsLTECell[f"G{i}"].value) in range(700, 710):
                    n_cells += 1
                cell_id = get_parameter_value(proj_paramsLTECell[f"F{i}"].value, 2)
                if cell_id not in cell_sectors:
                    cell_sectors.append(cell_id)
                me_info.initialized = True
                # me_info.cells.append(Cell(proj_paramsLTECell[f"K{i}"].value,  # cell_id
                #                           # proj_paramsLTECell[f"H{i}"].value,  # frequency
                #                           proj_paramsLTECell[f"H{i}"].value[-2:-1],  # sector_id
                #                           proj_paramsLTECell[f"R{i}"].value,  # TAC
                #                           proj_paramsLTECell[f"O{i}"].value))  # PCI
        if not hasattr(me_info, "enbcu"):
            raise Exception("Unable to find " + me_id + " in CUEUtranCellFDDLTE")
        me_info.n_cells = n_cells
        me_info.sectors = sorted(cell_sectors)
        mes_data.append(me_info)

        # for i in range(6, get_max_row(target_me, sheet_nameGNBCU) + 1):
        #     if proj_paramsGNBCU[f"D{i}"].value == me_id:
        #         if not proj_paramsGNBCU[f"F{i}"].value.startswith("214-09"):
        #             me_info.gnodeb_id = find_cucp(proj_paramsGNBCU[f"F{i}"].value)
        #             me_info.mcc = proj_paramsGNBCU[f"N{i}"].value.split("-")[0]
        #             me_info.mnc = proj_paramsGNBCU[f"N{i}"].value.split("-")[1]  # has a leading 0 for values under 10
    return target_me, mes_data


def generate_cosites(working_wb, mes_data):
    sheet = "EUtranRelationFDDLTE"
    sheet_data = working_wb[sheet]

    total_relations = 0
    initial_max_row = get_max_row(working_wb, sheet) + 1

    for me in mes_data:
        curr_me_relations = 0
        print("-----", me.me_id, "-----")
        target_sectors = [str(sector) for sector in me.sectors if int(sector) in range(90, 97)]
        source_sectors = list(set(me.sectors) - set(target_sectors))
        source_sectors = sorted(source_sectors)

        target_row = get_max_row(working_wb, sheet) + 1

        global_idx = 0

        for s_sector in source_sectors:
            idx, new_rel = find_max_cell_rel_idx(target_row, sheet_data, me, s_sector)
            if global_idx == 0:
                global_idx = idx
            for t_sector in target_sectors:
                print(s_sector, "--", t_sector)
                rel_exists = relation_exists(s_sector, t_sector, sheet_data, initial_max_row, me)
                if rel_exists < 0:
                    copy_row(idx, target_row, working_wb, sheet, sheet_data)
                    replace_4g_cosite_params(target_row, sheet_data, me, t_sector)
                    replace_relation_params(target_row, sheet_data, new_rel)
                    set_user_label(target_row, sheet_data, me)
                    configure_ca(target_row, sheet_data)
                    add_modifier(target_row, "A", sheet_data)
                    target_row += 1
                    new_rel += 1
                    curr_me_relations += 1
                else:
                    print("Relation", s_sector, "->", t_sector, "already exists")
                    set_user_label(rel_exists, sheet_data, me)
                    configure_ca(rel_exists, sheet_data)
                    add_modifier(rel_exists, "M", sheet_data)

        print("------")
        for t_sector in target_sectors:
            _, new_rel = find_max_cell_rel_idx(target_row, sheet_data, me, t_sector)
            for s_sector in [*source_sectors, *[ib_sector for ib_sector in target_sectors if ib_sector != t_sector]]:
                print(t_sector, "--", s_sector)
                rel_exists = relation_exists(t_sector, s_sector, sheet_data, initial_max_row, me)
                if rel_exists < 0:
                    copy_row(global_idx, target_row, working_wb, sheet, sheet_data)
                    replace_4g_cosite_params(target_row, sheet_data, me, s_sector)
                    replace_relation_params(target_row, sheet_data, new_rel, t_sector)
                    set_user_label(target_row, sheet_data, me)
                    configure_ca(target_row, sheet_data)
                    add_modifier(target_row, "A", sheet_data)
                    target_row += 1
                    new_rel += 1
                    curr_me_relations += 1
                else:
                    print("Relation", t_sector, "->", s_sector, "already exists")
                    set_user_label(rel_exists, sheet_data, me)
                    configure_ca(rel_exists, sheet_data)
                    add_modifier(rel_exists, "M", sheet_data)

        print("Added", curr_me_relations, "new relations for", me.me_id)
        total_relations += curr_me_relations
    print("Added", total_relations, "total relations")


if __name__ == "__main__":
    generate_4g_4g_cosite("PVAX0443", None, True)
