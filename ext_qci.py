from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime as dt
from ME_Data import ME, Cell
from utils import *


def generate_extqci(me_id, workbook=None, save_file=False):
    mes = [me.strip().upper() for me in me_id.split(";")]
    target_me, mes_data = load_input(mes, workbook)

    generate_EnDCDRXCycle(target_me, mes_data)  # 4 records/cell
    generate_NSADRX(target_me, mes_data)  # 2 records/cell
    generate_EnDCPBR(target_me, mes_data)  # 12 records/cell
    generate_NsaNgbrMLCPProfile(target_me, mes_data)  # 4 records/cell
    generate_NsaLcgProfile(target_me, mes_data)  # 1 records/cell

    if save_file:
        if len(mes) == 1:
            target_me.save(f"data/{mes_data[0].site_name}_singleSite_Baseline_{dt.now().strftime('%d%m%Y_%H%M%S')}.xlsx")
        else:
            target_me.save(f"data/{mes_data[0].site_name}_multiSite_Baseline_{dt.now().strftime('%d%m%Y_%H%M%S')}.xlsx")
    else:
        return target_me


def load_input(me_list, target):
    mes_data = []
    if target is None:
        target_me = load_workbook("data/input.xlsx")
    else:
        target_me = target
    sheet_nameNRCell = "NRCellDU"
    try:
        proj_paramsNRCell = target_me[sheet_nameNRCell]
    except:
        raise Exception("Cannot find " + sheet_nameNRCell + " MO in input file")

    for me_id in me_list:
        me_info = ME(me_id)
        n_cells = 0
        for i in range(6, get_max_row(target_me, sheet_nameNRCell) + 1):
            if proj_paramsNRCell[f"D{i}"].value == me_id:
                if not me_info.initialized:
                    me_info.subnetid = proj_paramsNRCell[f"C{i}"].value
                    me_info.gnbdu = find_cucp(proj_paramsNRCell[f"F{i}"].value)
                    me_info.site_name = proj_paramsNRCell[f"E{i}"].value
                    # me_info.user_label = proj_paramsNRCell[f"H{i}"].value
                    # me_info.cells = []
                if int(proj_paramsNRCell[f"G{i}"].value) in range(700, 710):
                    n_cells += 1
                me_info.initialized = True
                # me_info.cells.append(Cell(proj_paramsNRCell[f"K{i}"].value,  # cell_id
                #                           # proj_paramsNRCell[f"H{i}"].value,  # frequency
                #                           proj_paramsNRCell[f"H{i}"].value[-2:-1],  # sector_id
                #                           proj_paramsNRCell[f"R{i}"].value,  # TAC
                #                           proj_paramsNRCell[f"O{i}"].value))  # PCI
        if not hasattr(me_info, "gnbdu"):
            raise Exception("Unable to find " + me_id + " in NRCellDU")
        me_info.n_cells = n_cells
        mes_data.append(me_info)

        # for i in range(6, get_max_row(target_me, sheet_nameGNBCU) + 1):
        #     if proj_paramsGNBCU[f"D{i}"].value == me_id:
        #         if not proj_paramsGNBCU[f"F{i}"].value.startswith("214-09"):
        #             me_info.gnodeb_id = find_cucp(proj_paramsGNBCU[f"F{i}"].value)
        #             me_info.mcc = proj_paramsGNBCU[f"N{i}"].value.split("-")[0]
        #             me_info.mnc = proj_paramsGNBCU[f"N{i}"].value.split("-")[1]  # has a leading 0 for values under 10
    return target_me, mes_data


def generate_EnDCDRXCycle(working_wb, mes_data):
    sheet = "EnDCDRXCycle"
    sheet_data = working_wb[sheet]

    source_rows = []
    counted_values = []
    initial_last_row = target_row = get_max_row(working_wb, sheet) + 1
    for row in range(6, target_row):
        if len(source_rows) == 4:
            break
        if (sheet_data[f"G{row}"].value == "17" and sheet_data[f"H{row}"].value == "130") or \
                (sheet_data[f"G{row}"].value == "18" and sheet_data[f"H{row}"].value == "130") or \
                (sheet_data[f"G{row}"].value == "19" and sheet_data[f"H{row}"].value == "140") or \
                (sheet_data[f"G{row}"].value == "20" and sheet_data[f"H{row}"].value == "140"):
            if sheet_data[f"G{row}"].value not in counted_values:
                counted_values.append(sheet_data[f"G{row}"].value)
                source_rows.append(row)

    if not source_rows:
        raise Exception("No QCI 130 nor 140 has been found in", sheet)
    elif len(source_rows) < 4:
        raise Exception("Not found all values for QCI 130 or 140 in", sheet)

    for me in mes_data:
        for cell in range(me.n_cells):
            for i in range(len(source_rows)):
                if not exists_qci_relation(source_rows[i], sheet_data, int(cell) + 700, initial_last_row, me, "H"):
                    copy_row(source_rows[i], target_row, working_wb, sheet, sheet_data)
                    replace_site_cell_id(target_row, sheet_data, me.gnbdu, int(cell) + 700, me)
                    add_modifier(target_row, "A", sheet_data)
                    target_row += 1


def generate_NSADRX(working_wb, mes_data):
    sheet = "NSADRX"
    sheet_data = working_wb[sheet]

    source_rows = []
    counted_values = []
    initial_last_row = target_row = get_max_row(working_wb, sheet) + 1
    for row in range(6, target_row):
        if len(source_rows) == 2:
            break
        if (sheet_data[f"G{row}"].value == "10" and "800" in sheet_data[f"F{row}"].value.split(",")[1]) or \
                (sheet_data[f"G{row}"].value == "11" and "800" in sheet_data[f"F{row}"].value.split(",")[1]):
            if sheet_data[f"G{row}"].value not in counted_values:
                counted_values.append(sheet_data[f"G{row}"].value)
                source_rows.append(row)

    if len(source_rows) < 2:
        for row in range(6, target_row):
            if len(source_rows) == 2:
                break
            if (sheet_data[f"G{row}"].value == "10" and "500" in sheet_data[f"F{row}"].value.split(",")[1]) or \
                    (sheet_data[f"G{row}"].value == "11" and "500" in sheet_data[f"F{row}"].value.split(",")[1]):
                if sheet_data[f"G{row}"].value not in counted_values:
                    counted_values.append(sheet_data[f"G{row}"].value)
                    source_rows.append(row)

    if not source_rows:
        raise Exception("No 500 nor 800 data has been found in", sheet)
    elif len(source_rows) < 2:
        raise Exception("Not found all values for 500 or 800 in", sheet)

    for me in mes_data:
        for cell in range(me.n_cells):
            for i in range(len(source_rows)):
                if not exists_qci_relation(source_rows[i], sheet_data, int(cell) + 700, initial_last_row, me, "N"):
                    copy_row(source_rows[i], target_row, working_wb, sheet, sheet_data)
                    replace_site_cell_id(target_row, sheet_data, me.gnbdu, int(cell) + 700, me)
                    sheet_data[f"I{target_row}"].value = "80"
                    add_modifier(target_row, "A", sheet_data)
                    target_row += 1


def generate_EnDCPBR(working_wb, mes_data):
    sheet = "EnDCPBR"
    sheet_data = working_wb[sheet]

    source_rows = []
    counted_values = []
    initial_last_row = target_row = get_max_row(working_wb, sheet) + 1
    for row in range(6, target_row):
        if len(source_rows) == 12:
            break
        if (sheet_data[f"G{row}"].value == "31" and sheet_data[f"H{row}"].value in ["130", "140"]) and "800" in sheet_data[f"F{row}"].value.split(",")[1] or \
                (sheet_data[f"G{row}"].value == "32" and sheet_data[f"H{row}"].value in ["130", "140"]) and "800" in sheet_data[f"F{row}"].value.split(",")[1] or \
                (sheet_data[f"G{row}"].value == "33" and sheet_data[f"H{row}"].value in ["130", "140"]) and "800" in sheet_data[f"F{row}"].value.split(",")[1] or \
                (sheet_data[f"G{row}"].value == "34" and sheet_data[f"H{row}"].value in ["130", "140"]) and "800" in sheet_data[f"F{row}"].value.split(",")[1] or \
                (sheet_data[f"G{row}"].value == "35" and sheet_data[f"H{row}"].value in ["130", "140"]) and "800" in sheet_data[f"F{row}"].value.split(",")[1] or \
                (sheet_data[f"G{row}"].value == "36" and sheet_data[f"H{row}"].value in ["130", "140"]) and "800" in sheet_data[f"F{row}"].value.split(",")[1] or \
                (sheet_data[f"G{row}"].value == "37" and sheet_data[f"H{row}"].value in ["130", "140"]) and "800" in sheet_data[f"F{row}"].value.split(",")[1] or \
                (sheet_data[f"G{row}"].value == "38" and sheet_data[f"H{row}"].value in ["130", "140"]) and "800" in sheet_data[f"F{row}"].value.split(",")[1] or \
                (sheet_data[f"G{row}"].value == "39" and sheet_data[f"H{row}"].value in ["130", "140"]) and "800" in sheet_data[f"F{row}"].value.split(",")[1] or \
                (sheet_data[f"G{row}"].value == "40" and sheet_data[f"H{row}"].value in ["130", "140"]) and "800" in sheet_data[f"F{row}"].value.split(",")[1] or \
                (sheet_data[f"G{row}"].value == "41" and sheet_data[f"H{row}"].value in ["130", "140"]) and "800" in sheet_data[f"F{row}"].value.split(",")[1] or \
                (sheet_data[f"G{row}"].value == "42" and sheet_data[f"H{row}"].value in ["130", "140"]) and "800" in sheet_data[f"F{row}"].value.split(",")[1]:
            if sheet_data[f"G{row}"].value not in counted_values:
                counted_values.append(sheet_data[f"G{row}"].value)
                source_rows.append(row)

    if len(source_rows) < 12:
        for row in range(6, target_row):
            if len(source_rows) == 12:
                break
            if (sheet_data[f"G{row}"].value == "31" and sheet_data[f"H{row}"].value in ["130", "140"]) and "500" in \
                    sheet_data[f"F{row}"].value.split(",")[1] or \
                    (sheet_data[f"G{row}"].value == "32" and sheet_data[f"H{row}"].value in ["130", "140"]) and "500" in \
                    sheet_data[f"F{row}"].value.split(",")[1] or \
                    (sheet_data[f"G{row}"].value == "33" and sheet_data[f"H{row}"].value in ["130", "140"]) and "500" in \
                    sheet_data[f"F{row}"].value.split(",")[1] or \
                    (sheet_data[f"G{row}"].value == "34" and sheet_data[f"H{row}"].value in ["130", "140"]) and "500" in \
                    sheet_data[f"F{row}"].value.split(",")[1] or \
                    (sheet_data[f"G{row}"].value == "35" and sheet_data[f"H{row}"].value in ["130", "140"]) and "500" in \
                    sheet_data[f"F{row}"].value.split(",")[1] or \
                    (sheet_data[f"G{row}"].value == "36" and sheet_data[f"H{row}"].value in ["130", "140"]) and "500" in \
                    sheet_data[f"F{row}"].value.split(",")[1] or \
                    (sheet_data[f"G{row}"].value == "37" and sheet_data[f"H{row}"].value in ["130", "140"]) and "500" in \
                    sheet_data[f"F{row}"].value.split(",")[1] or \
                    (sheet_data[f"G{row}"].value == "38" and sheet_data[f"H{row}"].value in ["130", "140"]) and "500" in \
                    sheet_data[f"F{row}"].value.split(",")[1] or \
                    (sheet_data[f"G{row}"].value == "39" and sheet_data[f"H{row}"].value in ["130", "140"]) and "500" in \
                    sheet_data[f"F{row}"].value.split(",")[1] or \
                    (sheet_data[f"G{row}"].value == "40" and sheet_data[f"H{row}"].value in ["130", "140"]) and "500" in \
                    sheet_data[f"F{row}"].value.split(",")[1] or \
                    (sheet_data[f"G{row}"].value == "41" and sheet_data[f"H{row}"].value in ["130", "140"]) and "500" in \
                    sheet_data[f"F{row}"].value.split(",")[1] or \
                    (sheet_data[f"G{row}"].value == "42" and sheet_data[f"H{row}"].value in ["130", "140"]) and "500" in \
                    sheet_data[f"F{row}"].value.split(",")[1]:
                if sheet_data[f"G{row}"].value not in counted_values:
                    counted_values.append(sheet_data[f"G{row}"].value)
                    source_rows.append(row)

    if not source_rows:
        raise Exception("No 500 nor 800 data has been found in", sheet)
    elif len(source_rows) < 12:
        raise Exception("Not found all values for 500 or 800 in", sheet)

    for me in mes_data:
        for cell in range(me.n_cells):
            for i in range(len(source_rows)):
                if not exists_qci_relation(source_rows[i], sheet_data, int(cell) + 700, initial_last_row, me, "H"):
                    copy_row(source_rows[i], target_row, working_wb, sheet, sheet_data)
                    replace_site_cell_id(target_row, sheet_data, me.gnbdu, int(cell) + 700, me)
                    add_modifier(target_row, "A", sheet_data)
                    target_row += 1


def generate_NsaNgbrMLCPProfile(working_wb, mes_data):
    sheet = "NsaNgbrMLCPProfile"
    sheet_data = working_wb[sheet]

    source_rows = []
    counted_values = []
    initial_last_row = target_row = get_max_row(working_wb, sheet) + 1
    for row in range(6, target_row):
        if len(source_rows) == 4:
            break
        if (sheet_data[f"G{row}"].value == "9" and "800" in sheet_data[f"F{row}"].value.split(",")[1]) or \
                (sheet_data[f"G{row}"].value == "10" and "800" in sheet_data[f"F{row}"].value.split(",")[1]) or \
                (sheet_data[f"G{row}"].value == "11" and "800" in sheet_data[f"F{row}"].value.split(",")[1]) or \
                (sheet_data[f"G{row}"].value == "12" and "800" in sheet_data[f"F{row}"].value.split(",")[1]):
            if sheet_data[f"G{row}"].value not in counted_values:
                counted_values.append(sheet_data[f"G{row}"].value)
                source_rows.append(row)

    if len(source_rows) < 4:
        for row in range(6, target_row):
            if len(source_rows) == 4:
                break
            if (sheet_data[f"G{row}"].value == "9" and "500" in sheet_data[f"F{row}"].value.split(",")[1]) or \
                    (sheet_data[f"G{row}"].value == "10" and "500" in sheet_data[f"F{row}"].value.split(",")[1]) or \
                    (sheet_data[f"G{row}"].value == "11" and "500" in sheet_data[f"F{row}"].value.split(",")[1]) or \
                    (sheet_data[f"G{row}"].value == "12" and "500" in sheet_data[f"F{row}"].value.split(",")[1]):
                if sheet_data[f"G{row}"].value not in counted_values:
                    counted_values.append(sheet_data[f"G{row}"].value)
                    source_rows.append(row)

    if not source_rows:
        raise Exception("No 500 nor 800 data has been found in", sheet)
    elif len(source_rows) < 4:
        raise Exception("Not found all values for 500 or 800 in", sheet)

    for me in mes_data:
        for cell in range(me.n_cells):
            for i in range(len(source_rows)):
                if not exists_qci_relation(source_rows[i], sheet_data, int(cell) + 700, initial_last_row, me, "J"):
                    copy_row(source_rows[i], target_row, working_wb, sheet, sheet_data)
                    replace_site_cell_id(target_row, sheet_data, me.gnbdu, int(cell) + 700, me)
                    add_modifier(target_row, "A", sheet_data)
                    target_row += 1


def generate_NsaLcgProfile(working_wb, mes_data):
    sheet = "NsaLcgProfile"
    sheet_data = working_wb[sheet]

    source_rows = []
    counted_values = []
    initial_last_row = target_row = get_max_row(working_wb, sheet) + 1
    for row in range(6, target_row):
        if len(source_rows) == 1:
            break
        if sheet_data[f"G{row}"].value == "1" and "800" in sheet_data[f"F{row}"].value.split(",")[1]:
            if sheet_data[f"G{row}"].value not in counted_values:
                counted_values.append(sheet_data[f"G{row}"].value)
                source_rows.append(row)

    if len(source_rows) < 1:
        for row in range(6, target_row):
            if len(source_rows) == 1:
                break
            if sheet_data[f"G{row}"].value == "1" and "500" in sheet_data[f"F{row}"].value.split(",")[1]:
                if sheet_data[f"G{row}"].value not in counted_values:
                    counted_values.append(sheet_data[f"G{row}"].value)
                    source_rows.append(row)

    if not source_rows:
        raise Exception("No 500 nor 800 data has been found in", sheet)
    elif len(source_rows) < 1:
        raise Exception("Not found all values for 500 or 800 in", sheet)

    for me in mes_data:
        for cell in range(me.n_cells):
            for i in range(len(source_rows)):
                if not exists_qci_relation(source_rows[i], sheet_data, int(cell) + 700, initial_last_row, me, "J", "NsaLcgProfile"):
                    copy_row(source_rows[i], target_row, working_wb, sheet, sheet_data)
                    replace_site_cell_id(target_row, sheet_data, me.gnbdu, int(cell) + 700, me)
                    replace_drb_qci_list(target_row, sheet_data)
                    add_modifier(target_row, "M", sheet_data)
                    target_row += 1


if __name__ == "__main__":
    generate_extqci("PVAX0443", None, True)