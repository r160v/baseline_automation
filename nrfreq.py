from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime as dt
from ME_Data import ME, Cell
from utils import *


def generate_nrfreqs(me_id, workbook=None, save_file=False):
    mes = [me.strip().upper() for me in me_id.split(";")]
    target_me, mes_data = load_input(mes)
    generate_nrfreqparafddlte(target_me, mes_data)
    generate_frequencybandlist(target_me, mes_data)
    generate_nrfreq(target_me, mes_data)

    if save_file:
        if len(mes) == 1:
            target_me.save(f"data/{mes_data[0].site_name}_singleSite_Baseline_{dt.now().strftime('%d%m%Y_%H%M%S')}.xlsx")
        else:
            target_me.save(f"data/{mes_data[0].site_name}_multiSite_Baseline_{dt.now().strftime('%d%m%Y_%H%M%S')}.xlsx")
    else:
        return target_me


def generate_nrfreqparafddlte(working_wb, mes_data):
    sheet = "NRFreqParaFDDLTE"
    sheet_data = working_wb[sheet]

    initial_max_row = target_row = get_max_row(working_wb, sheet) + 1

    for me in mes_data:
        for cell in me.cells:
            moid = str(int(cell.cell_id) + 200)
            if not exists_moid(sheet_data, initial_max_row, moid, me):
                copy_row(6, target_row, working_wb, sheet, sheet_data)
                replace_nrfreqparafddlte(target_row, sheet_data, me, moid, cell.cell_id)
                add_modifier(target_row, "A", sheet_data)
                target_row += 1


def generate_frequencybandlist(working_wb, mes_data):
    sheet = "FrequencyBandList"
    sheet_data = working_wb[sheet]

    initial_max_row = target_row = get_max_row(working_wb, sheet) + 1
    moid = "3"

    for me in mes_data:
        if not exists_moid_freqband(sheet_data, initial_max_row, moid, me):
            copy_row(6, target_row, working_wb, sheet, sheet_data)
            replace_freqbandlist(target_row, sheet_data, me, moid)
            add_modifier(target_row, "A", sheet_data)
            target_row += 1


def generate_nrfreq(working_wb, mes_data):
    sheet = "NRFreq"
    sheet_data = working_wb[sheet]

    initial_max_row = target_row = get_max_row(working_wb, sheet) + 1
    moid = "3"

    for me in mes_data:
        if not exists_moid_nrfreq(sheet_data, initial_max_row, moid, me):
            copy_row(6, target_row, working_wb, sheet, sheet_data)
            replace_nrfreq(target_row, sheet_data, me, moid)
            add_modifier(target_row, "A", sheet_data)
            target_row += 1


def load_input(me_list):
    mes_data = []
    target_me = load_workbook("data/input.xlsx")
    sheet_nameCUE = "CUEUtranCellFDDLTE"
    sheet_nameGNBCU = "GNBCUCPFunction"
    try:
        proj_paramsCUE = target_me[sheet_nameCUE]
    except:
        raise Exception("Cannot find " + sheet_nameCUE + " MO in input file")
    try:
        proj_paramsGNBCU = target_me[sheet_nameGNBCU]
    except:
        raise Exception("Cannot find " + sheet_nameGNBCU + " MO in input file")

    for me_id in me_list:
        me_info = ME(me_id)
        for i in range(6, get_max_row(target_me, sheet_nameCUE) + 1):
            if proj_paramsCUE[f"D{i}"].value == me_id:
                if not me_info.initialized:
                    me_info.subnetid = proj_paramsCUE[f"C{i}"].value
                    me_info.nodeb_id = find_cucp(proj_paramsCUE[f"F{i}"].value)
                    me_info.site_name = proj_paramsCUE[f"E{i}"].value
                    me_info.user_label = proj_paramsCUE[f"H{i}"].value
                    me_info.cells = []
                    me_info.initialized = True
                if int(proj_paramsCUE[f"K{i}"].value) not in [*range(30, 34), *range(90, 94)]:
                    me_info.cells.append(Cell(proj_paramsCUE[f"K{i}"].value,  # cell_id
                                              proj_paramsCUE[f"H{i}"].value[-2:-1],  # sector_id
                                              proj_paramsCUE[f"R{i}"].value,  # TAC
                                              proj_paramsCUE[f"O{i}"].value))  # PCI

        if not hasattr(me_info, "nodeb_id"):
            raise Exception("Unable to find " + me_id + " in CUEUtranCellFDDLTE")

        for i in range(6, get_max_row(target_me, sheet_nameGNBCU) + 1):
            if proj_paramsGNBCU[f"D{i}"].value == me_id:
                if not proj_paramsGNBCU[f"F{i}"].value.startswith("214-09"):
                    me_info.gnodeb_id = find_cucp(proj_paramsGNBCU[f"F{i}"].value)
                    me_info.mcc = proj_paramsGNBCU[f"N{i}"].value.split("-")[0]
                    me_info.mnc = proj_paramsGNBCU[f"N{i}"].value.split("-")[1]  # has a leading 0 for values under 10

        if not hasattr(me_info, "gnodeb_id"):
            raise Exception("Unable to find " + me_id + " in GNBCUCPFunction")
        mes_data.append(me_info)
    return target_me, mes_data


if __name__ == "__main__":
    generate_nrfreqs("PVAX0443;PVAX0815", None, True)
