from openpyxl.utils import get_column_letter

id_cell = {"0": "L1A", "1": "L2A", "2": "L3A", "3": "L4A", "10": "N1A", "11": "N2A", "12": "N3A", "13": "N4A",
           "20": "M1A", "21": "M2A", "22": "M3A", "23": "M4A", "30": "T1A", "31": "T2A", "32": "T3A", "33": "T4A",
           "40": "L1B", "41": "L2B", "42": "L3B", "43": "L4B", "90": "Y1A", "91": "Y2A", "92": "Y3A", "93": "Y4A"}


def get_min_row(workbook, sht_name):
    return workbook[sht_name].min_row


def get_max_row(workbook, sht_name):
    return workbook[sht_name].max_row_nobreak


def get_min_col(workbook, sht_name):
    return workbook[sht_name].min_column


def get_max_col(workbook, sht_name):
    return workbook[sht_name].max_column


def find_cucp(cell_value):
    cell_conts = cell_value.split(",")
    cucp = cell_conts[0]
    init = cucp.rfind("=")
    return cucp[init + 1:]


def get_parameter_value(cell_value, idx):
    cell_conts = cell_value.split(",")
    cell_id = cell_conts[idx]
    init = cell_id.rfind("=")
    return cell_id[init + 1:]


def find_max_cell_rel_idx(target_row, bl_data, me, s_sector):
    max_eutran_rel = 0
    max_idx = 0
    for i in range(6, target_row):
        if bl_data[f"D{i}"].value == me.me_id and get_parameter_value(bl_data[f"F{i}"].value, 2) == s_sector:
            relation = get_parameter_value(bl_data[f"F{i}"].value, 4)
            try:
                eutran_rel = int(relation)
                if eutran_rel > max_eutran_rel:
                    max_eutran_rel = eutran_rel
                    max_idx = i
            except:
                continue

    return max_idx, max_eutran_rel + 1


def set_user_label(target_row, bl_data, me):
    cell_id = get_parameter_value(bl_data[f"F{target_row}"].value, 2)
    if cell_id in id_cell:
        bl_data[f"H{target_row}"].value = me.me_id + id_cell[cell_id]


def configure_ca(target_row, bl_data):
    source = get_parameter_value(bl_data[f"F{target_row}"].value, 2)
    target = get_parameter_value(bl_data[f"J{target_row}"].value, 2)

    if source[-1] == target[-1] and ((int(source) not in range(90, 95)) or (int(source) in range(90, 95) and
                                                                            (int(source) - 70 == int(target)))):
        bl_data[f"Q{target_row}"].value = "Same Coverage[1]"
    else:
        bl_data[f"Q{target_row}"].value = "Neighboring[0]"

    if source[-1] == target[-1]:
        bl_data[f"AC{target_row}"].value = "UL&DL CA Support[12]"
    elif abs(int(source) - int(target)) <= 3:
        bl_data[f"AC{target_row}"].value = "Only UL Comp Support[1]"
    else:
        bl_data[f"AC{target_row}"].value = "Not Support[0]"


def copy_row(source, target, workbook, sheet, bl_data):
    max_col = get_max_col(workbook, sheet)
    for col in range(1, max_col + 1):
        bl_data[f"{get_column_letter(col)}{target}"].value = \
            bl_data[f"{get_column_letter(col)}{source}"].value


def add_modifier(target_row, modifier, bl_data):
    bl_data[f"A{target_row}"].value = modifier


def find_reference_row(sheet_data, last_row, me):
    for i in range(6, last_row + 1):
        if sheet_data[f"D{i}"].value == me.me_id:
            return i


def replace_site_cell_id(target_row, bl_data, site_id, cell_id, me):
    bl_data[f"C{target_row}"].value = me.subnetid
    bl_data[f"D{target_row}"].value = me.me_id
    bl_data[f"E{target_row}"].value = me.me_id
    ldn = bl_data[f"F{target_row}"].value.split(",")
    curr_site = ldn[0]
    curr_cell = ldn[1]
    init = curr_site.rfind("=")
    ldn[0] = ldn[0].replace(ldn[0][init + 1:len(ldn[0])], str(site_id))
    init = curr_cell.rfind("=")
    ldn[1] = ldn[1].replace(ldn[1][init + 1:len(ldn[1])], str(cell_id))
    bl_data[f"F{target_row}"].value = ",".join(ldn)


def replace_drb_qci_list(target_row, bl_data):
    ldn = bl_data[f"F{target_row}"].value.split(",")
    lcg_profile = ldn[4]
    init = lcg_profile.rfind("=")
    ldn[4] = ldn[4].replace(ldn[4][init + 1:len(ldn[4])], "7")
    bl_data[f"F{target_row}"].value = ",".join(ldn)
    bl_data[f"G{target_row}"].value = "7"
    bl_data[f"H{target_row}"].value = "7"
    bl_data[f"I{target_row}"].value = "7;8;9;130;140"


def replace_relation_params(target_row, bl_data, new_rel, cell_id=None):
    bl_data[f"G{target_row}"].value = str(new_rel)
    ldn = bl_data[f"F{target_row}"].value.split(",")
    relation = ldn[4]
    init = relation.rfind("=")
    ldn[4] = ldn[4].replace(ldn[4][init + 1:len(ldn[4])], str(new_rel))
    if cell_id is not None:
        tcell_id = ldn[2]
        init = tcell_id.rfind("=")
        ldn[2] = ldn[2].replace(ldn[2][init + 1:len(ldn[2])], str(cell_id))
    bl_data[f"F{target_row}"].value = ",".join(ldn)


def replace_4g_cosite_params(target_row, sheet_data, me, t_sector):
    sheet_data[f"J{target_row}"].value = "ENBCUCPFunction=" + me.enbcu + ",CULTE=1,CUEUtranCellFDDLTE=" + t_sector
    sheet_data[f"L{target_row}"].value = ""
    sheet_data[f"N{target_row}"].value = "Yes[1]"
    sheet_data[f"O{target_row}"].value = "Yes[1]"
    sheet_data[f"P{target_row}"].value = "Yes[1]"


def replace_nrfreqparafddlte(target_row, sheet_data, me, moid, cell_id):
    sheet_data[f"C{target_row}"].value = me.subnetid
    sheet_data[f"D{target_row}"].value = me.me_id
    sheet_data[f"E{target_row}"].value = me.me_id
    sheet_data[f"F{target_row}"].value = "ENBCUCPFunction=" + me.nodeb_id + ",CULTE=1,CUEUtranCellFDDLTE=" + cell_id + \
                                         ",EUtranCellMeasFDDLTE=1,NRFreqParaFDDLTE=" + moid

    sheet_data[f"G{target_row}"].value = moid
    sheet_data[f"H{target_row}"].value = moid
    sheet_data[f"I{target_row}"].value = "28"
    sheet_data[f"J{target_row}"].value = "763.45"
    sheet_data[f"O{target_row}"].value = "245"
    sheet_data[f"X{target_row}"].value = "0001"


def replace_freqbandlist(target_row, sheet_data, me, moid):
    sheet_data[f"C{target_row}"].value = me.subnetid
    sheet_data[f"D{target_row}"].value = me.me_id
    sheet_data[f"E{target_row}"].value = me.me_id
    sheet_data[f"F{target_row}"].value = "GNBCUCPFunction=" + me.gnodeb_id + ",NRFreq=3,FrequencyBandList=" + moid
    sheet_data[f"G{target_row}"].value = moid
    sheet_data[f"H{target_row}"].value = "28"


def replace_nrfreq(target_row, sheet_data, me, moid):
    sheet_data[f"C{target_row}"].value = me.subnetid
    sheet_data[f"D{target_row}"].value = me.me_id
    sheet_data[f"E{target_row}"].value = me.me_id
    sheet_data[f"F{target_row}"].value = "GNBCUCPFunction=" + me.gnodeb_id + ",NRFreq=" + moid
    sheet_data[f"G{target_row}"].value = moid
    sheet_data[f"H{target_row}"].value = "763.45"


def exists_moid(sheet_data, last_row, candidate_moid, me):
    for i in range(6, last_row):
        if sheet_data[f"D{i}"].value != me.me_id:
            continue
        if sheet_data[f"G{i}"].value == candidate_moid:
            return True
    return False


def exists_qci_relation(row, sheet_data, cell_id, last_row, me, qci_col, sheet=None):
    moid = sheet_data[f"G{row}"].value
    qci = sheet_data[f"{qci_col}{row}"].value
    for i in range(6, last_row):
        if sheet_data[f"D{i}"].value != me.me_id:
            continue
        if sheet is not None:
            if sheet_data[f"G{i}"].value == "7" and sheet_data[f"H{i}"].value == "7" and \
                    sheet_data[f"I{i}"].value == "7;8;9;130;140" and \
                    get_parameter_value(sheet_data[f"F{i}"].value, 1) == str(cell_id):
                return True
        else:
            if sheet_data[f"G{i}"].value == moid and sheet_data[f"{qci_col}{i}"].value == qci and \
                    get_parameter_value(sheet_data[f"F{i}"].value, 1) == str(cell_id):
                return True
    return False





def exists_moid_freqband(sheet_data, last_row, candidate_moid, me):
    for i in range(6, last_row):
        if sheet_data[f"D{i}"].value != me.me_id:
            continue
        if sheet_data[f"G{i}"].value == candidate_moid and sheet_data[f"H{i}"].value == "28":
            return True
    return False


def exists_moid_nrfreq(sheet_data, last_row, candidate_moid, me):
    for i in range(6, last_row):
        if sheet_data[f"D{i}"].value != me.me_id:
            continue
        if sheet_data[f"G{i}"].value == candidate_moid and sheet_data[f"H{i}"].value == "763.45":
            return True
    return False


def get_moid(sheet_data, last_row, candidate_moid):
    local_cand = candidate_moid
    found = False

    while not found:
        for i in range(6, last_row + 1):
            if sheet_data[f"G{i}"].value == local_cand:
                rel = candidate_moid.split("-")
                try:
                    rel[2] = str(int(rel[2]) + 1)
                    local_cand = "-".join(rel)
                    break
                except:
                    raise TypeError("Relation including not only numbers")
        found = True
    return local_cand


def replace_5g_cosite_params(target_row, sheet_data, me, comb, last_row):
    sheet_data[f"C{target_row}"].value = me.subnetid
    sheet_data[f"D{target_row}"].value = me.me_id
    sheet_data[f"E{target_row}"].value = me.me_id
    cell_relation = get_parameter_value(sheet_data[f"F{target_row}"].value, 2)
    new_relation = cell_relation[:-3] + comb[1]

    new_relation = get_moid(sheet_data, last_row, new_relation)

    sheet_data[f"F{target_row}"].value = "GNBCUCPFunction=" + me.gnbdu + ",NRCellCU=" + comb[0] + \
                                         ",NRCellRelation=" + new_relation
    sheet_data[f"G{target_row}"].value = new_relation
    sheet_data[f"H{target_row}"].value = ""
    sheet_data[f"I{target_row}"].value = "GNBCUCPFunction=" + me.gnbdu + ",NRCellCU=" + comb[1]


def relation_exists(s_sector, t_sector, sheet_data, last_row, me):
    for i in range(6, last_row):
        if sheet_data[f"J{i}"].value == "" or sheet_data[f"D{i}"].value != me.me_id:
            continue
        csource = get_parameter_value(sheet_data[f"F{i}"].value, 2)
        ctarget = get_parameter_value(sheet_data[f"J{i}"].value, 2)
        if csource == s_sector and ctarget == t_sector and sheet_data[f"L{i}"].value == "":
            return i
    return -1


def relation_exists_5g(comb, sheet_data, last_row, me):
    for i in range(6, last_row + 1):
        if sheet_data[f"H{i}"].value != "" or sheet_data[f"D{i}"].value != me.me_id:
            continue
        csource = get_parameter_value(sheet_data[f"F{i}"].value, 1)
        ctarget = get_parameter_value(sheet_data[f"I{i}"].value, 1)
        if csource == comb[0] and ctarget == comb[1]:
            return i
    return -1
